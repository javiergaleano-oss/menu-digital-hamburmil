from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash
import pandas as pd
import os
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

app = Flask(__name__)
app.secret_key = "supersecretkey"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(BASE_DIR, "BASE DE DATOS MENU COMIDAS RAPIDAS.csv")


# ==============================
# NUMERO DE PEDIDO
# ==============================
def generar_numero_pedido():
    archivo = os.path.join(BASE_DIR, "contador_pedidos.json")
    hoy = datetime.now().strftime("%Y-%m-%d")

    if not os.path.exists(archivo):
        data = {"fecha": hoy, "contador": 1}
    else:
        with open(archivo, "r") as f:
            data = json.load(f)

        if data["fecha"] != hoy:
            data["contador"] = 1
            data["fecha"] = hoy
        else:
            data["contador"] += 1

    with open(archivo, "w") as f:
        json.dump(data, f)

    return data["contador"]


# ==============================
# EXCEL LIMPIO
# ==============================
def guardar_ventas_excel(carrito, pago_efectivo, pago_nequi):

    archivo = os.path.join(BASE_DIR, "ventas_diarias.xlsx")
    fecha = datetime.now().strftime("%Y-%m-%d")

    datos = {}
    efectivo = 0
    nequi = 0

    if os.path.exists(archivo):
        try:
            df_existente = pd.read_excel(archivo, sheet_name=fecha)

            for _, row in df_existente.iterrows():
                if pd.isna(row["REFERENCIA"]):
                    continue

                ref = str(row["REFERENCIA"])
                datos[ref] = {
                    "producto": row["PRODUCTO"],
                    "cantidad": row["CANTIDAD"],
                    "precio": row["PRECIO UNITARIO"],
                    "total": row["TOTAL"]
                }

            wb = load_workbook(archivo)
            ws = wb[fecha]

            for row in range(2, ws.max_row + 1):
                if ws[f"D{row}"].value == "EFECTIVO":
                    efectivo = ws[f"E{row}"].value or 0
                if ws[f"D{row}"].value == "NEQUI":
                    nequi = ws[f"E{row}"].value or 0

        except:
            pass

    for item in carrito:
        ref = item["REFERENCIA"]

        if ref not in datos:
            datos[ref] = {
                "producto": item["DESCRIPCION"],
                "cantidad": 0,
                "precio": item["PRECIO"],
                "total": 0
            }

        datos[ref]["cantidad"] += 1
        datos[ref]["total"] += item["PRECIO"]

    efectivo += pago_efectivo
    nequi += pago_nequi

    filas = []
    for ref, d in datos.items():
        filas.append({
            "PRODUCTO": d["producto"],
            "REFERENCIA": ref,
            "CANTIDAD": d["cantidad"],
            "PRECIO UNITARIO": d["precio"],
            "TOTAL": d["total"]
        })

    df_final = pd.DataFrame(filas)

    with pd.ExcelWriter(archivo, engine="openpyxl", mode="w") as writer:
        df_final.to_excel(writer, sheet_name=fecha, index=False)

    wb = load_workbook(archivo)
    ws = wb[fecha]

    bold = Font(bold=True)

    # LIMPIAR PAGOS
    filas_a_eliminar = []
    for row in range(2, ws.max_row + 1):
        if ws[f"D{row}"].value in ["EFECTIVO", "NEQUI", "TOTAL PAGOS"]:
            filas_a_eliminar.append(row)

    for row in reversed(filas_a_eliminar):
        ws.delete_rows(row)

    # TOTAL GENERAL
    ultima_fila = ws.max_row + 1
    fila_fin = len(df_final) + 1

    ws[f"D{ultima_fila}"] = "TOTAL GENERAL"
    ws[f"E{ultima_fila}"] = f"=SUM(E2:E{fila_fin})"

    # PAGOS
    fila_pagos = ultima_fila + 2

    ws[f"D{fila_pagos}"] = "EFECTIVO"
    ws[f"E{fila_pagos}"] = efectivo

    ws[f"D{fila_pagos+1}"] = "NEQUI"
    ws[f"E{fila_pagos+1}"] = nequi

    ws[f"D{fila_pagos+2}"] = "TOTAL PAGOS"
    ws[f"E{fila_pagos+2}"] = efectivo + nequi

    for i in [ultima_fila, fila_pagos, fila_pagos+1, fila_pagos+2]:
        ws[f"D{i}"].font = bold
        ws[f"E{i}"].font = bold

    wb.save(archivo)


# ==============================
# MENU
# ==============================
def cargar_menu():
    df = pd.read_csv(CSV_PATH, encoding="utf-8", sep=";")
    df.columns = df.columns.str.strip().str.upper().str.replace(" ", "", regex=False)
    df["REFERENCIA"] = df["REFERENCIA"].astype(str)
    df["PRECIO"] = pd.to_numeric(df["PRECIO"], errors="coerce").fillna(0)
    return df


@app.context_processor
def carrito_global():
    carrito = session.get("carrito", [])
    total = sum(item["PRECIO"] for item in carrito)
    return dict(carrito_cantidad=len(carrito), carrito_total=total)


# ==============================
# RUTAS
# ==============================
@app.route("/")
def index():
    menu = cargar_menu()
    categorias = sorted(menu["CATEGORIA"].dropna().unique())
    return render_template("categorias.html", categorias=categorias)


@app.route("/categoria/<nombre>")
def ver_categoria(nombre):
    menu = cargar_menu()
    productos = menu[menu["CATEGORIA"] == nombre]
    return render_template("productos.html", categoria=nombre, productos=productos.to_dict(orient="records"))


@app.route("/agregar", methods=["POST"])
def agregar():
    if "carrito" not in session:
        session["carrito"] = []

    session["carrito"].append({
        "REFERENCIA": request.form.get("referencia"),
        "DESCRIPCION": request.form.get("descripcion"),
        "PRECIO": float(request.form.get("precio")),
        "CATEGORIA": request.form.get("categoria"),
        "SALSAS": ", ".join(request.form.getlist("salsas")),
        "EXTRAS": ", ".join(request.form.getlist("extras"))
    })

    return redirect(url_for("ver_carrito"))


@app.route("/eliminar/<int:index>")
def eliminar(index):
    carrito = session.get("carrito", [])
    if 0 <= index < len(carrito):
        carrito.pop(index)
    session["carrito"] = carrito
    return redirect(url_for("ver_carrito"))


@app.route("/carrito")
def ver_carrito():
    carrito = session.get("carrito", [])
    total = sum(item["PRECIO"] for item in carrito)
    return render_template("carrito.html", carrito=carrito, total=total)


@app.route("/finalizar", methods=["POST"])
def finalizar():

    carrito = session.get("carrito", [])
    total = sum(item["PRECIO"] for item in carrito)

    pago_efectivo = float(request.form.get("pago_efectivo") or 0)
    pago_nequi = float(request.form.get("pago_nequi") or 0)

    total_pagado = pago_efectivo + pago_nequi

    if total_pagado < total:
        flash("❌ Pago insuficiente", "error")
        return redirect(url_for("ver_carrito"))

    cambio = total_pagado - total

    numero_pedido = generar_numero_pedido()

    pedido = {
        "numero": numero_pedido,
        "nombre": request.form.get("nombre"),
        "tipo_entrega": request.form.get("tipo_entrega"),
        "direccion": request.form.get("direccion"),
        "mesa": request.form.get("mesa"),
        "efectivo": pago_efectivo,
        "nequi": pago_nequi,
        "cambio": cambio
    }

    session["pedido"] = pedido

    guardar_ventas_excel(carrito, pago_efectivo, pago_nequi)

    return render_template("confirmacion.html", pedido=pedido, carrito=carrito, total=total)


@app.route("/ticket")
def ticket():
    carrito = session.get("carrito", [])
    pedido = session.get("pedido", {})
    total = sum(item["PRECIO"] for item in carrito)
    return render_template("ticket.html", carrito=carrito, total=total, pedido=pedido)


@app.route("/descargar_excel")
def descargar_excel():
    archivo = os.path.join(BASE_DIR, "ventas_diarias.xlsx")
    if os.path.exists(archivo):
        return send_file(archivo, as_attachment=True)
    return "No hay archivo"


@app.route("/limpiar")
def limpiar():
    session.clear()
    return redirect(url_for("index"))


if __name__ == "__main__":
    app.run(debug=True)